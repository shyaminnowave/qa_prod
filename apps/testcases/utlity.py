from abc import ABCMeta, abstractmethod
from openpyxl import load_workbook
from apps.testcases.models import (
    TestCaseModel,
    TestReport,
    NatcoStatus,
    TestCaseChoices,
    TestCaseStep,
)
from apps.stbs.models import NactoManufacturesLanguage, STBNodeConfig, NatcoRelease
from analytiqa.helpers.renders import ResponseInfo
from rest_framework import status
from django.db import transaction
from functools import lru_cache


class ExcelFileFactory:

    def __init__(self, file, **kwargs):
        self.response_format = ResponseInfo().response
        self.file = file
        super().__init__(**kwargs)

    def _init_workbook(self):
        wb = load_workbook(self.file)
        return wb.active


class ReportExcel(ExcelFileFactory):

    def __init__(self, **kwargs):
        super(ReportExcel, self).__init__(**kwargs)

    def _init_workbook(self):
        self.ws = super(ReportExcel, self)._init_workbook()
        return self.ws

    def get_testcase(self, jira_id):
        queryset = TestCaseModel.objects.get(jira_id=jira_id)
        return queryset

    def get_node(self, node_id):
        queryset = NatcoRelease.objects.get(id=node_id)
        return queryset

    def import_data(self):
        print("Inside the Instance")
        _test_result = []
        try:
            for row in self._init_workbook().iter_rows(min_row=2, values_only=True):
                print(
                    f"jira-id: {row[4]}, load: {row[5]},  cpu: {row[6]}, ram: {row[7]}, cpu_75: {row[16]}, ram_76: {row[17]}, load_75: {row[15]}, date: {row[2]}, starttime: {row[8]}, endtime: {row[9]}"
                )
                if row[4]:
                    _data = {
                        "run_type": row[0],
                        "date": str(row[2]),
                        "iteration_number": int(row[3]),
                        "testcase": self.get_testcase(row[4]),
                        "loadtime": float(row[5]),
                        "cpu_hundred_percentile": float(row[6]),
                        "ram_hundred_percentile": float(row[7]),
                        "start_time": str(row[8]),
                        "end_time": str(row[9]),
                        "job_id": row[10],
                        "node": self.get_node(row[11]),
                        "loadtime_percentile": row[15],
                        "cpu_usage_percentile": row[16],
                        "ram_usage_percentile": row[17],
                        "failure_reason": row[12],
                        "result": row[13],
                        "comment": row[13],
                    }
                    print(_data)
                    _test_result.append(TestReport(**_data))
                else:
                    break
            with transaction.atomic():
                TestReport.objects.bulk_create(_test_result)
        except Exception as e:
            self.response_format["status"] = False
            self.response_format["status_code"] = status.HTTP_400_BAD_REQUEST
            self.response_format["data"] = "Error"
            self.response_format["massage"] = str(e)
            return self.response_format
        self.response_format["status"] = True
        self.response_format["status_code"] = status.HTTP_200_OK
        self.response_format["data"] = "Success"
        self.response_format["message"] = "TestCase Uploaded Successfully"
        return self.response_format


class TestCaseExcel(ExcelFileFactory):

    def __init__(self, **kwargs):
        super(TestCaseExcel, self).__init__(**kwargs)

    def _init_workbook(self):
        self.ws = super(TestCaseExcel, self)._init_workbook()
        return self.ws

    def import_data(self):
        test_case = None
        _data = dict()
        _step_data = dict()
        testcase_list = []
        step_list = []
        natco_list = []
        natco = NactoManufactureLanguage.objects.all()
        try:
            for row in self._init_workbook().iter_rows(min_row=2, values_only=True):
                if row[2] is not None:
                    jira_id_parts = str(row[2]).split("-")
                    _data = {
                        "jira_id": jira_id_parts[-1],
                        "jira_summary": row[7],
                        "test_description": row[7],
                        "test_name": row[6],
                        "testcase_type": TestCaseChoices.SMOKE,
                    }
                    testcase_list.append(TestCaseModel(**_data))
                    test_case = jira_id_parts[-1]
                    for data in natco:
                        natco_list.append(
                            NatcoStatus(
                                natco=data.natco,
                                language=data.language_name,
                                device=data.device_name,
                                test_case_id=test_case,
                            )
                        )
                    if row[2] and row[8]:
                        _step_data = {
                            "testcase_id": test_case,
                            "step_id": int(row[8]),
                            "step_action": row[9],
                            "step_data": "",
                            "excepted_result": row[10],
                        }
                    step_list.append(TestCaseStep(**_step_data))
                elif row[2] is None and row[5] is not None:
                    _step_data = {
                        "testcase_id": test_case,
                        "step_id": int(row[8]),
                        "step_action": row[9],
                        "step_data": "",
                        "excepted_result": row[10],
                    }
                    step_list.append(TestCaseStep(**_step_data))
                elif row[2] is None and row[5] is None:
                    pass
            with transaction.atomic():
                TestCaseModel.objects.bulk_create(testcase_list)
                TestCaseStep.objects.bulk_create(step_list)
                NatcoStatus.objects.bulk_create(natco_list)
        except Exception as e:
            self.response_format["status"] = False
            self.response_format["status_code"] = status.HTTP_400_BAD_REQUEST
            # self.response_format['data'] = 'Error'
            self.response_format["message"] = str(e)
            return self.response_format
        self.response_format["status"] = True
        self.response_format["status_code"] = status.HTTP_200_OK
        self.response_format["data"] = "Success"
        self.response_format["message"] = "TestCase Uploaded Successfully"
        return self.response_format
